import base64
import hmac
from collections import Counter
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import requests
import streamlit as st
import streamlit.components.v1 as components


APP_DIR = Path(__file__).resolve().parent
LOCAL_DATA = APP_DIR / "TotoHistoryAll.xlsx"
GITHUB_OWNER = "wazley-hub"
GITHUB_REPO = "toto-predictor-lite"
GITHUB_BRANCH = "main"
GITHUB_DATA_PATH = "TotoHistoryAll.xlsx"
GITHUB_CONTENTS_API = (
    f"https://api.github.com/repos/{GITHUB_OWNER}/{GITHUB_REPO}/contents/"
    f"{GITHUB_DATA_PATH}"
)


st.set_page_config(page_title="Toto Predictor", page_icon="🎯", layout="wide")
st.markdown(
    """
    <style>
    .block-container {padding-top: 2rem; padding-bottom: 4rem; max-width: 1180px;}
    h1, h2, h3 {letter-spacing: -0.02em;}
    div[data-testid="stMetric"] {
        border: 1px solid #e5e7eb; border-radius: 12px; padding: 12px 16px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def pad4(value):
    try:
        if pd.isna(value):
            return "0000"
    except Exception:
        pass
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(4)[-4:]


def family4(value):
    return "".join(sorted(pad4(value)))


def normalize_history(df):
    aliases = {
        "DrawNo": "draw_no",
        "DrawDate": "draw_date",
        "1stPrizeNo": "first",
        "2ndPrizeNo": "second",
        "3rdPrizeNo": "third",
    }
    missing = [column for column in aliases if column not in df.columns]
    if missing:
        raise ValueError("Format data tidak lengkap.")
    out = df[list(aliases)].rename(columns=aliases).copy()
    for column in ("first", "second", "third"):
        out[column] = out[column].map(pad4)
    return out.reset_index(drop=True)


def get_secret(name):
    try:
        return str(st.secrets.get(name, "")).strip()
    except Exception:
        return ""


def github_headers(token=""):
    headers = {
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def fetch_github_workbook(token=""):
    response = requests.get(
        GITHUB_CONTENTS_API,
        params={"ref": GITHUB_BRANCH},
        headers=github_headers(token),
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    encoded = str(payload.get("content", "")).replace("\n", "")
    if not encoded:
        raise RuntimeError("Kandungan fail GitHub tidak dapat dibaca.")
    return base64.b64decode(encoded), str(payload.get("sha", ""))


@st.cache_data(ttl=60, show_spinner=False)
def load_history():
    try:
        workbook_bytes, _ = fetch_github_workbook(get_secret("LITE_GITHUB_TOKEN"))
        return (
            normalize_history(pd.read_excel(BytesIO(workbook_bytes))),
            "GitHub Toto Predictor Lite",
        )
    except Exception:
        if not LOCAL_DATA.exists():
            raise RuntimeError("Data GitHub tidak dapat dibaca dan fail fallback tiada.")
        return normalize_history(pd.read_excel(LOCAL_DATA)), "Fail fallback (read-only)"


def validate_update(raw_df, draw_no, draw_date, first, second, third):
    values = (first, second, third)
    if not draw_no.isdigit():
        return "Draw No mesti mengandungi nombor sahaja."
    if not draw_date.isdigit() or len(draw_date) != 8:
        return "Tarikh mesti dalam format YYYYMMDD, contohnya 20260725."
    try:
        datetime.strptime(draw_date, "%Y%m%d")
    except ValueError:
        return "Tarikh tidak sah."
    if not all(value.isdigit() and len(value) == 4 for value in values):
        return "1st, 2nd dan 3rd Prize mesti tepat empat digit."

    existing_draws = raw_df["DrawNo"].astype(str).str.replace(r"\.0$", "", regex=True)
    if draw_no in set(existing_draws):
        return f"Draw {draw_no} sudah wujud. Kemas kini dibatalkan."

    dates = pd.to_numeric(raw_df["DrawDate"], errors="coerce").dropna()
    if not dates.empty and int(draw_date) <= int(dates.max()):
        return "Tarikh baharu mesti selepas tarikh keputusan terakhir."
    return ""


def update_github_history(draw_no, draw_date, first, second, third):
    token = get_secret("LITE_GITHUB_TOKEN")
    if not token:
        raise RuntimeError("LITE_GITHUB_TOKEN belum ditetapkan dalam Streamlit Secrets.")

    workbook_bytes, current_sha = fetch_github_workbook(token)
    raw_df = pd.read_excel(BytesIO(workbook_bytes))
    required = ["DrawNo", "DrawDate", "1stPrizeNo", "2ndPrizeNo", "3rdPrizeNo"]
    if any(column not in raw_df.columns for column in required):
        raise RuntimeError("Format TotoHistoryAll.xlsx dalam repo tidak sah.")

    validation_error = validate_update(
        raw_df, draw_no, draw_date, first, second, third
    )
    if validation_error:
        raise ValueError(validation_error)

    new_row = {
        "DrawNo": int(draw_no),
        "DrawDate": int(draw_date),
        "1stPrizeNo": int(first),
        "2ndPrizeNo": int(second),
        "3rdPrizeNo": int(third),
    }
    updated = pd.concat([raw_df, pd.DataFrame([new_row])], ignore_index=True)
    output = BytesIO()
    updated.to_excel(output, index=False, engine="openpyxl")

    response = requests.put(
        GITHUB_CONTENTS_API,
        headers=github_headers(token),
        json={
            "message": f"Update Toto Predictor Lite Draw {draw_no}",
            "content": base64.b64encode(output.getvalue()).decode("ascii"),
            "sha": current_sha,
            "branch": GITHUB_BRANCH,
        },
        timeout=30,
    )
    response.raise_for_status()


def copy_button(label, value, key):
    import json

    payload = json.dumps(str(value))
    components.html(
        f"""
        <button onclick='navigator.clipboard.writeText({payload}).then(() => {{
            const msg = document.getElementById("msg_{key}");
            msg.innerText = "Disalin";
            setTimeout(() => msg.innerText = "", 1500);
        }})'
        style="border:0;border-radius:10px;background:#2563eb;color:white;
               padding:10px 16px;font-size:15px;font-weight:700;cursor:pointer;">
            {label}
        </button>
        <span id="msg_{key}" style="color:#15803d;font-size:14px;font-weight:600;margin-left:8px;"></span>
        """,
        height=48,
    )


def build_bridge_v1(first, second, third):
    numbers = [pad4(first), pad4(second), pad4(third)]
    existing = sorted(set("".join(numbers)))
    missing = sorted(set("0123456789") - set(existing))
    pair_rows = []
    base_pairs = []
    for source, number in zip(("1st", "2nd", "3rd"), numbers):
        for position, pair in zip(
            ("Front", "Middle", "Back"),
            (number[:2], number[1:3], number[2:4]),
        ):
            pair_rows.append(
                {"Source": source, "Pair Position": position, "Pair": pair}
            )
            base_pairs.append(pair)
    base_pairs = list(dict.fromkeys(base_pairs))

    rows = []
    seen = set()
    for pair in base_pairs:
        for missing_digit in missing:
            for existing_digit in existing:
                number = f"{pair}{missing_digit}{existing_digit}"
                family = family4(number)
                if family in seen:
                    continue
                seen.add(family)
                rows.append({"No": number, "Family": family, "Base Pair": pair})
    bridge_df = pd.DataFrame(rows, columns=["No", "Family", "Base Pair"])
    text = (
        "🧪 Toto Predictor - Bridge V1\n\n"
        f"Base Pairs:\n{' / '.join(base_pairs)}\n\n"
        f"Missing Digits:\n{' / '.join(missing)}\n\n"
        f"Existing Digits:\n{' / '.join(existing)}\n\n"
        f"Bridge Numbers (Total: {len(bridge_df)}):\n"
    )
    values = bridge_df["No"].tolist()
    text += "\n".join(
        " / ".join(values[index:index + 10])
        for index in range(0, len(values), 10)
    ) or "Tiada output."
    return pd.DataFrame(pair_rows), bridge_df, text


def build_bridge_v2(first, second, third):
    """Bridge V2: base pair + dua digit missing atau dua digit existing."""
    numbers = [pad4(first), pad4(second), pad4(third)]
    existing = sorted(set("".join(numbers)))
    missing = sorted(set("0123456789") - set(existing))
    pair_rows = []
    base_pairs = []
    for source, number in zip(("1st", "2nd", "3rd"), numbers):
        for position, pair in zip(
            ("Front", "Middle", "Back"),
            (number[:2], number[1:3], number[2:4]),
        ):
            pair_rows.append(
                {"Source": source, "Pair Position": position, "Pair": pair}
            )
            base_pairs.append(pair)
    base_pairs = list(dict.fromkeys(base_pairs))

    records = {}
    for pair in base_pairs:
        for pool, mode in ((missing, "2 Missing"), (existing, "2 Existing")):
            for digit_1 in pool:
                for digit_2 in pool:
                    if digit_1 == digit_2:
                        continue
                    number = f"{pair}{digit_1}{digit_2}"
                    record = records.setdefault(
                        number,
                        {"No": number, "Mode": set(), "Base Pair": set()},
                    )
                    record["Mode"].add(mode)
                    record["Base Pair"].add(pair)

    rows = [
        {
            "No": record["No"],
            "Mode": " / ".join(sorted(record["Mode"])),
            "Base Pair": " / ".join(sorted(record["Base Pair"])),
        }
        for record in records.values()
    ]
    bridge_df = pd.DataFrame(rows, columns=["No", "Mode", "Base Pair"])
    text = (
        "🧪 Toto Predictor - Bridge V2\n\n"
        f"Base Pairs:\n{' / '.join(base_pairs)}\n\n"
        f"Missing Digits:\n{' / '.join(missing)}\n\n"
        f"Existing Digits:\n{' / '.join(existing)}"
    )
    for mode in ("2 Missing", "2 Existing"):
        values = (
            bridge_df.loc[
                bridge_df["Mode"].str.contains(mode, regex=False), "No"
            ].tolist()
            if not bridge_df.empty
            else []
        )
        text += f"\n\n{mode} Numbers (Total: {len(values)}):\n"
        text += "\n".join(
            " / ".join(values[index:index + 10])
            for index in range(0, len(values), 10)
        ) or "Tiada output."
    return pd.DataFrame(pair_rows), bridge_df, text


@st.cache_data(show_spinner=False)
def run_lite_backtest(history, draw_count):
    """Uji Bridge V1 dan V2 sahaja pada peralihan draw sejarah."""
    test_history = history.tail(min(int(draw_count), len(history))).reset_index(drop=True)
    quick_rows = []
    detail_rows = []

    for index, source in test_history.iterrows():
        first = pad4(source["first"])
        second = pad4(source["second"])
        third = pad4(source["third"])
        pair_v1, bridge_v1, _ = build_bridge_v1(first, second, third)
        pair_v2, bridge_v2, _ = build_bridge_v2(first, second, third)

        has_next = index + 1 < len(test_history)
        next_row = test_history.iloc[index + 1] if has_next else None
        next_numbers = (
            [pad4(next_row[column]) for column in ("first", "second", "third")]
            if has_next
            else []
        )
        v1_families = set(bridge_v1["Family"].astype(str))
        v2_families = set(bridge_v2["No"].map(family4))
        v1_hits = [number for number in next_numbers if family4(number) in v1_families]
        v2_hits = [number for number in next_numbers if family4(number) in v2_families]

        source_result = " / ".join((first, second, third))
        next_result = " / ".join(next_numbers) if has_next else "Belum ada next draw"
        quick_rows.append(
            {
                "Source Draw": str(source["draw_no"]),
                "Source Result": source_result,
                "Next Draw": str(next_row["draw_no"]) if has_next else "Belum ada next draw",
                "Next Result": next_result,
                "Bridge V1 Hit No": " / ".join(v1_hits),
                "Bridge V2 Hit No": " / ".join(v2_hits),
            }
        )

        base_pairs = list(dict.fromkeys(pair_v1["Pair"].astype(str).tolist()))
        existing = sorted(set(first + second + third))
        missing = sorted(set("0123456789") - set(existing))
        detail_rows.append(
            {
                "Source Draw": str(source["draw_no"]),
                "Source Date": str(source["draw_date"]),
                "Source 1st": first,
                "Source 2nd": second,
                "Source 3rd": third,
                "Base Pairs": " / ".join(base_pairs),
                "Missing Digits": " / ".join(missing),
                "Existing Digits": " / ".join(existing),
                "Bridge V1 Count": len(bridge_v1),
                "Bridge V1 Numbers": " / ".join(bridge_v1["No"].astype(str)),
                "Bridge V2 Count": len(bridge_v2),
                "Bridge V2 Numbers": " / ".join(bridge_v2["No"].astype(str)),
                "Next Draw": str(next_row["draw_no"]) if has_next else "",
                "Next Result": " / ".join(next_numbers),
                "Bridge V1 Hit": "YES" if v1_hits else "NO",
                "Bridge V1 Hit No": " / ".join(v1_hits),
                "Bridge V2 Hit": "YES" if v2_hits else "NO",
                "Bridge V2 Hit No": " / ".join(v2_hits),
            }
        )

    quick_df = pd.DataFrame(quick_rows)
    detail_df = pd.DataFrame(detail_rows)
    completed = max(len(quick_df) - 1, 0)
    v1_hit_count = int(quick_df.iloc[:completed]["Bridge V1 Hit No"].ne("").sum())
    v2_hit_count = int(quick_df.iloc[:completed]["Bridge V2 Hit No"].ne("").sum())
    unique_hit_count = int(
        (
            quick_df.iloc[:completed]["Bridge V1 Hit No"].ne("")
            | quick_df.iloc[:completed]["Bridge V2 Hit No"].ne("")
        ).sum()
    )
    rate = lambda value: round((value / completed * 100), 1) if completed else 0.0
    summary_df = pd.DataFrame(
        {
            "Metric": [
                "Jumlah Draw", "Draw Selesai", "Draw Pending",
                "Bridge V1 Hit", "Bridge V1 Hit Rate %",
                "Bridge V2 Hit", "Bridge V2 Hit Rate %",
                "Bridge V1 atau V2 Hit", "Total Unique Hit Rate %",
            ],
            "Value": [
                len(quick_df), completed, len(quick_df) - completed,
                v1_hit_count, rate(v1_hit_count), v2_hit_count,
                rate(v2_hit_count), unique_hit_count, rate(unique_hit_count),
            ],
        }
    )
    return quick_df, summary_df, detail_df


@st.cache_data(show_spinner=False)
def build_lite_backtest_excel(quick_df, summary_df, detail_df):
    """Bina fail Excel Lite dengan tiga tab sahaja."""
    from openpyxl.styles import Alignment, Font, PatternFill

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        quick_df.to_excel(writer, sheet_name="Quick Review", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        detail_df.to_excel(writer, sheet_name="Detail", index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="17365D")
        header_font = Font(color="FFFFFF", bold=True)
        fills = {
            "v1": PatternFill("solid", fgColor="E2F0D9"),
            "v2": PatternFill("solid", fgColor="DDEBF7"),
        }

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[1].height = 32

        quick = workbook["Quick Review"]
        quick.sheet_properties.pageSetUpPr.fitToPage = True
        quick.page_setup.orientation = "landscape"
        quick.page_setup.fitToWidth = 1
        quick.column_dimensions["A"].width = 14
        quick.column_dimensions["B"].width = 24
        quick.column_dimensions["C"].width = 18
        quick.column_dimensions["D"].width = 24
        quick.column_dimensions["E"].width = 20
        quick.column_dimensions["F"].width = 20
        for row in range(2, quick.max_row + 1):
            quick.cell(row, 5).fill = fills["v1"]
            quick.cell(row, 6).fill = fills["v2"]
            for column in range(1, 7):
                quick.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)

        summary = workbook["Summary"]
        summary.column_dimensions["A"].width = 30
        summary.column_dimensions["B"].width = 18
        for row in range(2, summary.max_row + 1):
            summary.cell(row, 2).font = Font(bold=True)

        detail = workbook["Detail"]
        for column in detail.columns:
            letter = column[0].column_letter
            header = str(column[0].value)
            detail.column_dimensions[letter].width = 55 if "Numbers" in header else min(max(len(header) + 3, 13), 24)
        for row in range(2, detail.max_row + 1):
            for column in range(1, detail.max_column + 1):
                detail.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)
            detail.row_dimensions[row].height = 45

    output.seek(0)
    return output.getvalue()


def bridge_state(first, second, third):
    numbers = [pad4(first), pad4(second), pad4(third)]
    digit_counts = Counter("".join(numbers))
    unique_pairs = set()
    for number in numbers:
        unique_pairs.update((number[:2], number[1:3], number[2:4]))
    _, bridge_v1, _ = build_bridge_v1(*numbers)
    _, bridge_v2, _ = build_bridge_v2(*numbers)
    return {
        "missing": 10 - len(digit_counts),
        "max_repeat": max(digit_counts.values()),
        "repeat_digits": sum(value >= 2 for value in digit_counts.values()),
        "pairs": len(unique_pairs),
        "v1_count": len(bridge_v1),
        "v2_count": len(bridge_v2),
    }


@st.cache_data(show_spinner=False)
def choose_bridge_route(history, first, second, third, lookback=100):
    """Pilih tahap Bridge sahaja mengikut dasar Route Signal app utama."""
    current = bridge_state(first, second, third)
    samples = []
    start = max(0, len(history) - int(lookback) - 1)
    for index in range(start, len(history) - 1):
        source = history.iloc[index]
        target = history.iloc[index + 1]
        source_numbers = [pad4(source[column]) for column in ("first", "second", "third")]
        target_numbers = [pad4(target[column]) for column in ("first", "second", "third")]
        state = bridge_state(*source_numbers)
        _, v1_df, _ = build_bridge_v1(*source_numbers)
        _, v2_df, _ = build_bridge_v2(*source_numbers)
        target_families = {family4(number) for number in target_numbers}
        v1_hit = bool(set(v1_df["Family"].astype(str)) & target_families)
        v2_hit = bool(set(v2_df["No"].map(family4)) & target_families)
        distance = (
            abs(state["missing"] - current["missing"]) * 3
            + abs(state["max_repeat"] - current["max_repeat"]) * 2
            + abs(state["repeat_digits"] - current["repeat_digits"])
            + abs(state["pairs"] - current["pairs"])
            + abs(state["v1_count"] - current["v1_count"]) / 30
            + abs(state["v2_count"] - current["v2_count"]) / 60
        )
        samples.append(
            {
                "distance": float(distance),
                "recency": index,
                "v1_hit": v1_hit,
                "v2_hit": v2_hit,
            }
        )

    # App utama membandingkan lapan keadaan sejarah terdekat. Pada tahap Bridge,
    # V1 hanya dipilih apabila sokongannya mengatasi V2 dengan margin yang jelas;
    # selain itu pilihan kekal V2. Aras Match tidak dipaparkan dalam app Lite.
    nearest = sorted(samples, key=lambda row: (row["distance"], -row["recency"]))[:8]
    if not nearest:
        return {"route": "Bridge V2", "support": 0, "v1_rate": 0.0, "v2_rate": 0.0}
    weights = [1 / (1 + row["distance"]) for row in nearest]
    total_weight = sum(weights)
    v1_rate = sum(weight * row["v1_hit"] for weight, row in zip(weights, nearest)) / total_weight
    v2_rate = sum(weight * row["v2_hit"] for weight, row in zip(weights, nearest)) / total_weight
    route = "Bridge V1" if v1_rate > v2_rate + 0.25 else "Bridge V2"
    return {
        "route": route,
        "support": len(nearest),
        "v1_rate": round(v1_rate * 100, 1),
        "v2_rate": round(v2_rate * 100, 1),
    }


@st.cache_data(show_spinner=False)
def build_pair_priority(history, first, second, third):
    slots = [
        ("1st", "Front", 0, "first"),
        ("1st", "Middle", 1, "first"),
        ("1st", "Back", 2, "first"),
        ("2nd", "Front", 0, "second"),
        ("2nd", "Middle", 1, "second"),
        ("2nd", "Back", 2, "second"),
        ("3rd", "Front", 0, "third"),
        ("3rd", "Middle", 1, "third"),
        ("3rd", "Back", 2, "third"),
    ]
    v1_hits = Counter()
    transitions = len(history) - 1
    for index in range(transitions):
        source_numbers = [
            pad4(history.iloc[index][column])
            for column in ("first", "second", "third")
        ]
        existing = sorted(set("".join(source_numbers)))
        missing = sorted(set("0123456789") - set(existing))
        targets = {
            family4(history.iloc[index + 1][column])
            for column in ("first", "second", "third")
        }
        for source, position, start, column in slots:
            pair = pad4(history.iloc[index][column])[start:start + 2]
            v1_families = {
                family4(f"{pair}{missing_digit}{existing_digit}")
                for missing_digit in missing
                for existing_digit in existing
            }
            if v1_families & targets:
                v1_hits[(source, position)] += 1

    current = {"first": pad4(first), "second": pad4(second), "third": pad4(third)}
    rows = []
    for order, (source, position, start, column) in enumerate(slots):
        v1 = int(v1_hits[(source, position)])
        rows.append(
            {
                "Source": source,
                "Pair Position": position,
                "Current Pair": current[column][start:start + 2],
                "Historical Hit": v1,
                "Transitions": transitions,
                "_Order": order,
            }
        )
    ranked = pd.DataFrame(rows).sort_values(
        ["Historical Hit", "_Order"], ascending=[False, True], kind="stable"
    ).reset_index(drop=True)
    ranked.insert(0, "Priority", range(1, len(ranked) + 1))
    return ranked.drop(columns="_Order")


def build_pair_numbers(pair, audit_row, first, second, third):
    current = [pad4(first), pad4(second), pad4(third)]
    existing = sorted(set("".join(current)))
    missing = sorted(set("0123456789") - set(existing))
    pair = str(pair).zfill(2)[-2:]
    records = {}

    def add(number, route):
        family = family4(number)
        records.setdefault(
            (route, family),
            {"Pair": pair, "No": number, "Family": family, "Route": route},
        )

    for missing_digit in missing:
        for existing_digit in existing:
            add(f"{pair}{missing_digit}{existing_digit}", "Bridge V1")
    frame = pd.DataFrame(
        records.values(), columns=["Pair", "No", "Family", "Route"]
    )
    lines = [
        "🧭 Toto Predictor - Pilihan Mengikut Pair",
        "",
        f"Pair Pilihan: {pair}",
    ]
    values = frame["No"].tolist()
    lines.extend(["", f"Bridge V1 (Pilihan Unik: {len(values)}):"])
    lines.extend(
        " / ".join(values[index:index + 10])
        for index in range(0, len(values), 10)
    )
    return frame, "\n".join(lines)


def build_pair_numbers_v2(pair, first, second, third):
    current = [pad4(first), pad4(second), pad4(third)]
    existing = sorted(set("".join(current)))
    missing = sorted(set("0123456789") - set(existing))
    pair = str(pair).zfill(2)[-2:]
    records = {}
    for pool, mode in ((missing, "2 Missing"), (existing, "2 Existing")):
        for digit_1 in pool:
            for digit_2 in pool:
                if digit_1 == digit_2:
                    continue
                number = f"{pair}{digit_1}{digit_2}"
                record = records.setdefault(
                    number, {"Pair": pair, "No": number, "Mode": set()}
                )
                record["Mode"].add(mode)
    rows = [
        {
            "Pair": record["Pair"],
            "No": record["No"],
            "Mode": " / ".join(sorted(record["Mode"])),
        }
        for record in records.values()
    ]
    frame = pd.DataFrame(rows, columns=["Pair", "No", "Mode"])
    values = frame["No"].tolist()
    lines = [
        "🧭 Toto Predictor - Pilihan Pair Bridge V2",
        "",
        f"Pair Pilihan: {pair}",
        "",
        f"Bridge V2 (Pilihan Unik: {len(values)}):",
    ]
    lines.extend(
        " / ".join(values[index:index + 10])
        for index in range(0, len(values), 10)
    )
    return frame, "\n".join(lines)


def render_pair_selection(priority_df, first, second, third, version):
    st.subheader(f"🧭 Pilihan Pair {version}")
    st.caption(
        f"Buka pair yang dikehendaki untuk melihat dan menyalin nombor "
        f"{version} bagi pair itu sahaja."
    )
    ranking = " / ".join(dict.fromkeys(
        str(row["Current Pair"]).zfill(2)[-2:]
        for _, row in priority_df.iterrows()
    ))
    st.markdown(f"**Pair semasa:** {ranking}")

    shown = set()
    for _, audit_row in priority_df.iterrows():
        pair = str(audit_row["Current Pair"]).zfill(2)[-2:]
        if pair in shown:
            continue
        shown.add(pair)
        same_pair = priority_df[
            priority_df["Current Pair"].astype(str).str.zfill(2) == pair
        ]
        source_text = " / ".join(
            f'{row["Source"]} {row["Pair Position"]}'
            for _, row in same_pair.iterrows()
        )
        if version == "Bridge V1":
            numbers_df, copy_text = build_pair_numbers(
                pair, audit_row, first, second, third
            )
            copy_key = f"pair_v1_{pair}"
        else:
            numbers_df, copy_text = build_pair_numbers_v2(
                pair, first, second, third
            )
            copy_key = f"pair_v2_{pair}"
        with st.expander(
            f"Pair {pair} — {len(numbers_df)} pilihan", expanded=False
        ):
            st.caption(f"Sumber: {source_text}")
            copy_button(f"📋 Copy Pair {pair}", copy_text, copy_key)
            st.markdown(f"**{version} — {len(numbers_df)} pilihan unik**")
            st.dataframe(
                numbers_df.drop(columns=["Family", "Route"], errors="ignore"),
                hide_index=True,
                use_container_width=True,
            )


def build_chart_3d(first, second, third, bridge_v1_df, bridge_v2_df):
    numbers = [pad4(first), pad4(second), pad4(third)]
    digit_sums = [sum(int(digit) for digit in number) for number in numbers]
    digit_roots = [0 if value == 0 else 1 + (value - 1) % 9 for value in digit_sums]
    total_sum = str(sum(digit_sums))
    root_sum = str(sum(digit_roots))
    cross_rows = [
        "".join(str(int(top) + int(bottom)) for bottom in root_sum)
        for top in total_sum
    ]
    final_row = str(sum(int(digit) for digit in total_sum)) + str(
        sum(int(digit) for digit in root_sum)
    )
    derived_rows = cross_rows + [final_row]
    chart_rows = [total_sum, root_sum] + derived_rows

    choices = []
    seen = set()
    max_width = max(len(row) for row in derived_rows)
    for column in range(max_width):
        if all(column < len(row) for row in derived_rows):
            value = "".join(row[column] for row in derived_rows)
            if len(value) == 3 and ("Menegak", value) not in seen:
                seen.add(("Menegak", value))
                choices.append({"Pilihan": "Menegak", "3D": value})
    for row_index in range(len(derived_rows) - 1):
        top_row = derived_rows[row_index]
        bottom_row = derived_rows[row_index + 1]
        for column in range(min(len(top_row), len(bottom_row)) - 1):
            variants = [
                ("L Kiri", top_row[column] + bottom_row[column] + bottom_row[column + 1]),
                ("L Kanan", top_row[column + 1] + bottom_row[column + 1] + bottom_row[column]),
            ]
            if row_index < len(cross_rows) - 1:
                variants.append(
                    ("L Atas", top_row[column] + bottom_row[column] + top_row[column + 1])
                )
            for label, value in variants:
                if label != "Menegak" and any(
                    old_label != "Menegak" and old_value == value
                    for old_label, old_value in seen
                ):
                    continue
                if (label, value) not in seen:
                    seen.add((label, value))
                    choices.append({"Pilihan": label, "3D": value})
    choices_df = pd.DataFrame(choices, columns=["Pilihan", "3D"])

    confirmed = []
    for _, choice in choices_df.iterrows():
        anchor = str(choice["3D"])
        for bridge, frame in (("V1", bridge_v1_df), ("V2", bridge_v2_df)):
            for number in frame.get("No", pd.Series(dtype=str)).astype(str):
                if not (Counter(anchor) - Counter(pad4(number))):
                    confirmed.append(
                        {
                            "Pilihan": choice["Pilihan"],
                            "3D": anchor,
                            "No": pad4(number),
                            "Bridge": bridge,
                        }
                    )
    confirmed_df = pd.DataFrame(
        confirmed, columns=["Pilihan", "3D", "No", "Bridge"]
    ).drop_duplicates()
    vertical = choices_df.loc[choices_df["Pilihan"] == "Menegak", "3D"].tolist()
    l_choices = choices_df.loc[choices_df["Pilihan"] != "Menegak", "3D"].tolist()
    text = (
        "🧩 Toto Predictor - Carta 3D\n\n"
        f"Top 3: {' / '.join(numbers)}\n"
        f"Jumlah Digit: {' / '.join(str(value) for value in digit_sums)}\n"
        f"Digital Root: {' / '.join(str(value) for value in digit_roots)}\n"
        f"Asas: {total_sum} / {root_sum}\n\n"
        + "\n".join(chart_rows)
        + f"\n\nPilihan Menegak: {' / '.join(vertical) or 'Tiada'}"
        + f"\nPilihan L: {' / '.join(l_choices) or 'Tiada'}"
    )
    return text, chart_rows, vertical, l_choices, confirmed_df


st.title("🎯 Toto Predictor")
st.caption("Bridge Analysis Lite")

try:
    history, data_source = load_history()
except Exception as error:
    st.error(str(error))
    st.stop()

latest = history.iloc[-1]
st.subheader("📅 Keputusan Terbaru")
result_columns = st.columns(4)
result_columns[0].metric("Draw No", str(latest["draw_no"]))
result_columns[1].metric("1st Prize", pad4(latest["first"]))
result_columns[2].metric("2nd Prize", pad4(latest["second"]))
result_columns[3].metric("3rd Prize", pad4(latest["third"]))
st.caption(f'Tarikh: {latest["draw_date"]}')

with st.expander("📝 Update Keputusan", expanded=False):
    update_password = st.text_input(
        "Kata laluan kemas kini",
        type="password",
        key="lite_update_password_input",
    )
    update_columns = st.columns(2)
    new_draw_no = update_columns[0].text_input(
        "Draw No baharu", key="lite_new_draw_no"
    ).strip()
    new_draw_date = update_columns[1].text_input(
        "Tarikh baharu (YYYYMMDD)", key="lite_new_draw_date", max_chars=8
    ).strip()
    prize_columns = st.columns(3)
    new_first = prize_columns[0].text_input(
        "1st Prize baharu", key="lite_new_first", max_chars=4
    ).strip()
    new_second = prize_columns[1].text_input(
        "2nd Prize baharu", key="lite_new_second", max_chars=4
    ).strip()
    new_third = prize_columns[2].text_input(
        "3rd Prize baharu", key="lite_new_third", max_chars=4
    ).strip()
    confirm_update = st.checkbox(
        "Saya telah menyemak dan mengesahkan keputusan ini.",
        key="lite_confirm_update",
    )

    if st.button("Simpan Keputusan", key="lite_save_result"):
        configured_password = get_secret("LITE_UPDATE_PASSWORD")
        if not configured_password:
            st.error("LITE_UPDATE_PASSWORD belum ditetapkan dalam Streamlit Secrets.")
        elif not hmac.compare_digest(update_password, configured_password):
            st.error("Kata laluan kemas kini salah.")
        elif not confirm_update:
            st.error("Tandakan kotak pengesahan sebelum menyimpan.")
        else:
            try:
                with st.spinner("Menyimpan keputusan ke repo Lite..."):
                    update_github_history(
                        new_draw_no,
                        new_draw_date,
                        new_first,
                        new_second,
                        new_third,
                    )
                load_history.clear()
                st.success(f"Draw {new_draw_no} berjaya disimpan.")
                st.rerun()
            except ValueError as error:
                st.error(str(error))
            except requests.HTTPError as error:
                status = getattr(error.response, "status_code", "")
                st.error(
                    f"GitHub menolak kemas kini ({status}). "
                    "Semak token dan akses repo Lite."
                )
            except Exception as error:
                st.error(f"Kemas kini gagal: {error}")

st.divider()
with st.expander("🧪 Backtest Bridge V1 + V2", expanded=False):
    draw_options = [value for value in (30, 50, 100, 200, 500) if value <= len(history)]
    if len(history) not in draw_options and len(history) < 30:
        draw_options.append(len(history))
    default_index = draw_options.index(100) if 100 in draw_options else len(draw_options) - 1
    backtest_draws = st.selectbox(
        "Jumlah source draw untuk test",
        draw_options,
        index=default_index,
        key="lite_backtest_draws",
    )
    if st.button("Run Backtest", key="run_lite_backtest"):
        with st.spinner("Backtest sedang berjalan..."):
            quick_review, summary, detail = run_lite_backtest(
                history, backtest_draws
            )
            workbook_bytes = build_lite_backtest_excel(
                quick_review, summary, detail
            )
            st.session_state["lite_backtest_result"] = {
                "draws": backtest_draws,
                "quick": quick_review,
                "summary": summary,
                "detail": detail,
                "excel": workbook_bytes,
            }

    backtest_result = st.session_state.get("lite_backtest_result")
    if backtest_result:
        completed = int(
            backtest_result["summary"].loc[
                backtest_result["summary"]["Metric"] == "Draw Selesai", "Value"
            ].iloc[0]
        )
        st.success(f"Backtest selesai: {completed} draw lengkap diuji.")
        with st.expander("Lihat Summary", expanded=False):
            st.dataframe(
                backtest_result["summary"],
                hide_index=True,
                use_container_width=True,
            )
        with st.expander("Lihat Quick Review", expanded=False):
            st.dataframe(
                backtest_result["quick"],
                hide_index=True,
                use_container_width=True,
                height=360,
            )
        with st.expander("Lihat Detail", expanded=False):
            st.dataframe(
                backtest_result["detail"],
                hide_index=True,
                use_container_width=True,
                height=360,
            )
        st.download_button(
            "📥 Download Backtest Excel",
            data=backtest_result["excel"],
            file_name=(
                f"Toto_Predictor_Lite_Backtest_"
                f"{backtest_result['draws']}_Draw.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_lite_backtest",
        )

st.divider()
st.subheader("🎲 Generate")
input_columns = st.columns(3)
first = input_columns[0].text_input("1st Prize", value=pad4(latest["first"]), max_chars=4)
second = input_columns[1].text_input("2nd Prize", value=pad4(latest["second"]), max_chars=4)
third = input_columns[2].text_input("3rd Prize", value=pad4(latest["third"]), max_chars=4)

if st.button("Generate", type="primary"):
    if not all(value.isdigit() and len(value) == 4 for value in (first, second, third)):
        st.error("Masukkan tepat empat digit untuk setiap keputusan.")
    else:
        st.session_state["generated_values"] = (first, second, third)

if "generated_values" not in st.session_state:
    st.stop()

first, second, third = st.session_state["generated_values"]
priority_df = build_pair_priority(history, first, second, third)
pair_df, bridge_df, bridge_text = build_bridge_v1(first, second, third)
bridge_v2_pair_df, bridge_v2_df, bridge_v2_text = build_bridge_v2(
    first, second, third
)

st.divider()
st.subheader("🧭 Route Selection")
route_result = choose_bridge_route(history, first, second, third, lookback=100)
selected_route = route_result["route"]
st.markdown(f"**Bridge Pilihan:** {selected_route}")
if selected_route == "Bridge V1":
    selected_numbers = bridge_df["No"].astype(str).tolist()
    selected_text = bridge_text
elif selected_route == "Bridge V2":
    selected_numbers = bridge_v2_df["No"].astype(str).tolist()
    selected_text = bridge_v2_text
else:
    selected_numbers = []
    selected_text = ""

if selected_numbers:
    with st.expander(
        f"Lihat nombor {selected_route} ({len(selected_numbers)})",
        expanded=False,
    ):
        st.markdown(" / ".join(selected_numbers))
        copy_button("📋 Copy", selected_text, "copy_lite_route")

st.divider()
st.subheader("🧪 Bridge V1")
st.caption(f"Jumlah pilihan unik: {len(bridge_df)}")
copy_button("📋 Copy Bridge V1", bridge_text, "bridge_v1")
with st.expander("Lihat Detail Bridge V1", expanded=False):
    st.markdown("**Base Pair**")
    st.dataframe(pair_df, hide_index=True, use_container_width=True)
    st.markdown("**Senarai Bridge**")
    st.dataframe(
        bridge_df.drop(columns=["Family"], errors="ignore"),
        hide_index=True,
        use_container_width=True,
    )

render_pair_selection(priority_df, first, second, third, "Bridge V1")

st.divider()
st.subheader("🧪 Bridge V2")
missing_count = int(
    bridge_v2_df["Mode"].str.contains("2 Missing", regex=False).sum()
) if not bridge_v2_df.empty else 0
existing_count = int(
    bridge_v2_df["Mode"].str.contains("2 Existing", regex=False).sum()
) if not bridge_v2_df.empty else 0
st.caption(
    f"Jumlah pilihan unik: {len(bridge_v2_df)} | "
    f"2 Missing: {missing_count} | 2 Existing: {existing_count}"
)
copy_button("📋 Copy Bridge V2", bridge_v2_text, "bridge_v2")
with st.expander("Lihat Detail Bridge V2", expanded=False):
    st.markdown("**Base Pair**")
    st.dataframe(
        bridge_v2_pair_df, hide_index=True, use_container_width=True
    )
    st.markdown("**Senarai Bridge**")
    st.dataframe(bridge_v2_df, hide_index=True, use_container_width=True)

render_pair_selection(priority_df, first, second, third, "Bridge V2")

st.divider()
st.subheader("🧩 Carta 3D")
chart_text, chart_rows, vertical, l_choices, confirmed_df = build_chart_3d(
    first, second, third, bridge_df, bridge_v2_df
)
st.code("\n".join(chart_rows), language=None)
st.markdown(
    f"**Pilihan Menegak:** {' / '.join(vertical) or 'Tiada'}  \n"
    f"**Pilihan L:** {' / '.join(l_choices) or 'Tiada'}  \n"
    f"**Carta 3D + Bridge:** {len(confirmed_df)}"
)
copy_button("📋 Copy Carta 3D", chart_text, "chart_3d")
